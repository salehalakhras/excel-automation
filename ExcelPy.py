import openpyxl
from openpyxl.styles import PatternFill
from tkinter import *
from tkinter import messagebox

root = Tk()
root.title("Excel")
root.geometry('300x300')

file_path = Label(root, text="File name:")
file_path.pack()
file_path_entry = Entry(root)
file_path_entry.pack()


def click():
    name = file_path_entry.get()
    if name != '':
        change_excel(name)
        root.destroy()
    else:
        messagebox.showerror('Error', 'Please enter file name')


def change_excel(name):
    excel = openpyxl.load_workbook(name + '.xlsx')
    excel1 = openpyxl.Workbook()
    sheet = excel.active
    s0 = excel1.active
    excel1.remove(s0)
    sales_sheet = excel1.create_sheet('Sales Filter')
    prod_sheet = excel1.create_sheet('Production Filter')
    s1 = excel1.create_sheet('Sales')
    report = excel1.create_sheet('Main Report')
    s4 = excel1.create_sheet('Open Stock')
    stock_sheet = excel1.create_sheet('Open Stock Filter')
    s5 = excel1.create_sheet('Closing Stock')
    closing_sheet = excel1.create_sheet('Closing Stock Filter')
    s6 = excel1.create_sheet('AP-PO')
    s7 = excel1.create_sheet('GR-PO')
    red_fill = PatternFill(start_color='FFBCBC', end_color='FFBCBC', fill_type='solid')
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

    ## delete rows with depth greater than 2
    for i in range(sheet.max_row, 1, -1):
        if sheet.cell(row=i, column=7).value > 2:
            sheet.delete_rows(i, 1)

    ## copying and seperating sales and production
    sales_copy = False
    prod_copy = False
    first_row = True
    for row in sheet.iter_rows():
        if first_row:
            sales_sheet.append([cell.value for cell in row])
            prod_sheet.append([cell.value for cell in row])
            stock_sheet.append([cell.value for cell in row])
            closing_sheet.append([cell.value for cell in row])
            first_row = False
        if row[6].value == 1:
            if row[7].value and row[7].value.lower() == 'production':
                prod_sheet.append([cell.value for cell in row])
                stock_sheet.append([cell.value for cell in row])
                closing_sheet.append([cell.value for cell in row])
                prod_copy = True
                sales_copy = False
            elif row[7].value and row[7].value.lower() == 'sales':
                sales_sheet.append([cell.value for cell in row])
                prod_copy = False
                sales_copy = True

        elif sales_copy:
            sales_sheet.append([cell.value for cell in row])
        elif prod_copy:
            prod_sheet.append([cell.value for cell in row])
            stock_sheet.append([cell.value for cell in row])
            closing_sheet.append([cell.value for cell in row])

    ## formatting and filling 'I' col
    # Sales Sheet
    for r in range(2, sales_sheet.max_row+1):
        if sales_sheet.cell(row=r, column=7).value == 1:
            for i in range(1, 11):
                sales_sheet.cell(row=r, column=i).fill = blue_fill
            recipeRow = r
            sales_sheet.cell(row=r, column=9).value = '=IFERROR(VLOOKUP(A' + str(r) + ',Sales!$1:$1048576,3,0),0)'
        elif sales_sheet.cell(row=r, column=7).value == 2:
            for i in range(1, 11):
                sales_sheet.cell(row=r, column=i).fill = red_fill
            sales_sheet.cell(row=r, column=9).value = '=$D' + str(r) + '*' + '$I' + str(recipeRow)

    # Production Sheet
    for r in range(2, prod_sheet.max_row+1):
        if prod_sheet.cell(row=r, column=7).value == 1:
            for i in range(1, 11):
                prod_sheet.cell(row=r, column=i).fill = blue_fill
            recipeRowP = r
            prod_sheet.cell(row=r, column=9).value = '=SUMIF(\'Sales Filter\'!A:A,A' + str(r) + ',\'Sales Filter\'!I:I)'
        elif prod_sheet.cell(row=r, column=7).value == 2:
            for i in range(1, 11):
                prod_sheet.cell(row=r, column=i).fill = red_fill
            prod_sheet.cell(row=r, column=9).value = '=$D' + str(r) + '*' + '$I' + str(recipeRowP)

    # Stock Sheet
    for r in range(2, stock_sheet.max_row+1):
        if stock_sheet.cell(row=r, column=7).value == 1:
            for i in range(1, 11):
                stock_sheet.cell(row=r, column=i).fill = blue_fill
            recipeRowS = r
            stock_sheet.cell(row=r, column=9).value = '=IFERROR(VLOOKUP(A' + str(r) + ',\'Open Stock\'!$1:$1048576,5,0),0)'
        elif stock_sheet.cell(row=r, column=7).value == 2:
            for i in range(1, 11):
                stock_sheet.cell(row=r, column=i).fill = red_fill
            stock_sheet.cell(row=r, column=9).value = '=$D' + str(r) + '*' + '$I' + str(recipeRowS)

    # Closing Stock Sheet
    for r in range(2, closing_sheet.max_row+1):
        if closing_sheet.cell(row=r, column=7).value == 1:
            for i in range(1, 11):
                closing_sheet.cell(row=r, column=i).fill = blue_fill
            recipeRowS = r
            closing_sheet.cell(row=r, column=9).value = '=IFERROR(VLOOKUP(A' + str(r) + ',\'Closing Stock\'!$1:$1048576,5,0),0)'
        elif closing_sheet.cell(row=r, column=7).value == 2:
            for i in range(1, 11):
                closing_sheet.cell(row=r, column=i).fill = red_fill
            closing_sheet.cell(row=r, column=9).value = '=$D' + str(r) + '*' + '$I' + str(recipeRowS)

    ## Main Report
    report.cell(row=1, column=15).value = 'Open Stock'
    report.cell(row=1, column=18).value = 'Purchased'
    report.cell(row=1, column=21).value = 'Sales'
    report.cell(row=1, column=24).value = 'Closing Stock'
    report.merge_cells(start_row=1, start_column=15, end_row=1, end_column=17)
    report.merge_cells(start_row=1, start_column=18, end_row=1, end_column=20)
    report.merge_cells(start_row=1, start_column=21, end_row=1, end_column=23)
    report.merge_cells(start_row=1, start_column=24, end_row=1, end_column=26)

    # Open Stock
    report.cell(row=2, column=15).value = 'Recipe Items'
    report.cell(row=2, column=16).value = 'Consumed Items'
    report.cell(row=2, column=17).value = 'Total'
    for r in range(3,2000):
        report.cell(row=r, column=15).value = '=SUMIF(\'Open Stock Filter\'!A:A,A' + str(r) + ',\'Open Stock Filter\'!I:I)'
        report.cell(row=r, column=16).value = '=IFERROR(VLOOKUP(A' + str(r) + ',\'Open Stock\'!$1:$1048576,5,0),0)'
        report.cell(row=r, column=17).value = '=$P' + str(r) + '+' + '$O' + str(r)

    # Purchased
    report.cell(row=2, column=18).value = 'GR/PO'
    report.cell(row=2, column=19).value = 'AP/PO'
    report.cell(row=2, column=20).value = 'Total'
    for r in range(3,2000):
        report.cell(row=r, column=18).value = '=IFERROR(VLOOKUP(A' + str(r) + ',\'GR-PO\'!$1:$1048576,3,0),0)'
        report.cell(row=r, column=19).value = '=IFERROR(VLOOKUP(A' + str(r) + ',\'AP-PO\'!$1:$1048576,3,0),0)'
        report.cell(row=r, column=20).value = '=$S' + str(r) + '+' + '$R' + str(r)

    # Sales
    report.cell(row=2, column=21).value = 'Recipe Items'
    report.cell(row=2, column=22).value = 'Consumed Items'
    report.cell(row=2, column=23).value = 'Total'
    for r in range(3,2000):
        report.cell(row=r, column=21).value = '=SUMIF(\'Production Filter\'!A:A,A' + str(r) + ',\'Production Filter\'!I:I)'
        report.cell(row=r, column=22).value = '=SUMIF(\'Sales Filter\'!A:A,A' + str(r) + ',\'Sales Filter\'!I:I)'
        report.cell(row=r, column=23).value = '=$V' + str(r) + '+' + '$U' + str(r)

    # Closing Stock
    report.cell(row=2, column=24).value = 'Recipe Items'
    report.cell(row=2, column=25).value = 'Consumed Items'
    report.cell(row=2, column=26).value = 'Total'
    for r in range(3,2000):
        report.cell(row=r, column=24).value = '=SUMIF(\'Closing Stock Filter\'!A:A,A' + str(r) + ',\'Closing Stock Filter\'!I:I)'
        report.cell(row=r, column=25).value = '=IFERROR(VLOOKUP(A' + str(r) + ',\'Closing Stock\'!$1:$1048576,5,0),0)'
        report.cell(row=r, column=26).value = '=$Y' + str(r) + '+' + '$X' + str(r)

    # Variance
    report.cell(row=2, column=27).value = 'Variance'
    for r in range(3,2000):
        report.cell(row=r, column = 27).value = '=($Z' + str(r) + '+ $W' + str(r) + ') - ($T' + str(r) + '+ $Q' + str(r) + ')'


    # Hide irrelevant Sheets
    sales_sheet.sheet_state = 'hidden'
    prod_sheet.sheet_state = 'hidden'
    stock_sheet.sheet_state = 'hidden'
    closing_sheet.sheet_state = 'hidden'


    # Hide Report Data except Total
    for col in ['O','P','R','S','U','V','X','Y']:
        report.column_dimensions[col].hidden = True

    # change report sheet tab color
    report.sheet_properties.tabColor = 'FFDD00'

    ## save the file
    excel1.save('output.xlsx')


btn = Button(root, text='CLick Here', command=click)
btn.pack()

root.mainloop()
